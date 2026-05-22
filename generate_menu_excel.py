"""
generate_menu_excel.py
Generate Excel import file for Art Village menu.
Run: uv run generate_menu_excel.py
"""

from openpyxl import Workbook

wb = Workbook()

# ── Categories ──
ws_categories = wb.active
ws_categories.title = "Categories"
ws_categories.append(["categoryId", "name", "sortOrder", "enabled"])
categories = [
    ("cat-main", "主食", 10, True),
    ("cat-set", "套餐", 20, True),
    ("cat-salad", "涼拌", 30, True),
    ("cat-soup", "湯品", 40, True),
    ("cat-tea", "茶品", 50, True),
    ("cat-drink", "飲品", 60, True),
    ("cat-dessert", "甜點", 70, True),
]
for cat in categories:
    ws_categories.append(cat)

# ── OptionGroups ──
ws_option_groups = wb.create_sheet("OptionGroups")
ws_option_groups.append(["groupId", "name", "type", "required", "sortOrder", "enabled"])
option_groups = [
    ("opt-protein", "加料", "single", True, 10, True),
    ("opt-ice", "冰塊", "single", False, 20, True),
    ("opt-sweet", "甜度", "single", False, 30, True),
]
for og in option_groups:
    ws_option_groups.append(og)

# ── OptionItems ──
ws_option_items = wb.create_sheet("OptionItems")
ws_option_items.append(["optionItemId", "groupId", "name", "sortOrder", "enabled"])
option_items = [
    # 加料
    ("opt-protein-chicken", "opt-protein", "加雞", 10, True),
    ("opt-protein-shrimp", "opt-protein", "加蝦", 20, True),
    ("opt-protein-squid", "opt-protein", "加魷", 30, True),
    ("opt-protein-seafood", "opt-protein", "加海鮮", 40, True),
    # 冰塊
    ("opt-ice-normal", "opt-ice", "正常冰", 10, True),
    ("opt-ice-less", "opt-ice", "少冰", 20, True),
    ("opt-ice-none", "opt-ice", "無冰", 30, True),
    # 甜度
    ("opt-sweet-normal", "opt-sweet", "正常甜", 10, True),
    ("opt-sweet-less", "opt-sweet", "少甜", 20, True),
    ("opt-sweet-none", "opt-sweet", "無甜", 30, True),
]
for oi in option_items:
    ws_option_items.append(oi)

# ── Products ──
ws_products = wb.create_sheet("Products")
ws_products.append(["productId", "categoryId", "name", "description", "price", "soldOut", "sortOrder", "enabled", "imageUrl"])

# 主食 - each dish with all protein variants as separate products
# Format: (base_name, base_price, variants: [(suffix, price_add), ...])
main_dishes = [
    ("泰式炒麵", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("咖哩炒麵", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("泰式炒飯", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("咖哩炒飯", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("泰式船麵", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("乾咖哩", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("泰式蝦醬空心菜", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
    ("泰式三杯空心菜", 90, [("加雞", 20), ("加蝦", 40), ("加魷", 40), ("加海鮮", 60)]),
]

# 三杯牛肉 has different pricing
main_dishes.append(("泰式三杯牛肉", 150, [("加蝦", 20), ("加魷", 20), ("加海鮮", 40)]))

# Dishes without protein variants
main_dishes_no_variants = [
    ("泰式咖哩麵包雞", 130),
    ("泰式綠咖哩雞", 130),
    ("泰式紅咖哩雞", 130),
    ("泰式綠咖哩牛肉", 150),
    ("泰式紅咖哩牛肉", 150),
    ("泰式綠咖哩羊肉", 150),
    ("泰式紅咖哩羊肉", 150),
    ("泰式綠咖哩海鮮", 180),
    ("泰式紅咖哩海鮮", 180),
    ("泰式打拋豬", 110),
    ("泰式打拋雞", 110),
    ("泰式打拋牛肉", 150),
    ("泰式打拋羊肉", 150),
    ("泰式打拋海鮮", 150),
    ("泰式酸魚炸蝦", 150),
    ("泰式香烤魚", 150),
    ("泰式香烤雞", 130),
    ("泰式香烤羊肉", 150),
    ("泰式烤魷", 130),
    ("泰式烤雞腿", 130),
    ("泰式烤蝦", 150),
    ("泰式烤章魚", 150),
    ("泰式咖哩蟹", 280),
    ("泰式咖哩蝦", 180),
    ("泰式咖哩羊肉", 150),
    ("泰式咖哩章魚", 150),
]

sort_order = 10
products = []

# Add main dishes with variants
for dish in main_dishes:
    base_name, base_price, variants = dish
    # Base product (no protein)
    products.append((f"prod-{base_name.lower()[:10].replace(' ', '-')}", base_name, base_price, sort_order))
    sort_order += 10
    # Variant products
    for suffix, price_add in variants:
        variant_name = f"{base_name}({suffix})"
        products.append((f"prod-{base_name.lower()[:5]}-{suffix[:1]}", variant_name, base_price + price_add, sort_order))
        sort_order += 10

# Add main dishes without variants
for dish in main_dishes_no_variants:
    name, price = dish
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 套餐
set_meals = [
    ("綜合料理套餐", 150),
    ("泰式料理套餐", 150),
]
for name, price in set_meals:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 涼拌
salads = [
    ("泰式凉拌青瓜", 70),
    ("泰式凉拌番茄", 70),
    ("泰式凉拌高丽菜", 70),
    ("泰式凉拌小瓜", 70),
    ("泰式凉拌黃瓜", 70),
    ("泰式凉拌蕨菜", 70),
    ("泰式凉拌木耳", 70),
    ("泰式凉拌秋葵", 70),
]
for name, price in salads:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 湯品
soups = [
    ("泰式酸辣蝦湯", 90),
    ("泰式酸辣菇湯", 70),
]
for name, price in soups:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 茶品
teas = [
    ("泰式紅茶", 50),
    ("泰式奶茶", 50),
]
for name, price in teas:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 飲品
drinks = [
    ("檸檬冰茶", 40),
    ("檸檬红茶", 40),
    ("檸檬綠茶", 40),
    ("檸檬愛玉", 40),
    ("愛玉冰", 40),
    ("冬瓜茶", 40),
    ("酸梅", 40),
    ("冬瓜汁", 50),
    ("檸檬汁", 50),
    ("柳橙汁", 50),
    ("芭樂汁", 50),
    ("啤酒", 80),
    ("清酒", 150),
    ("紅酒", 200),
]
for name, price in drinks:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 甜點
desserts = [
    ("香椰冰淇淋", 60),
    ("香椰冰淇淋(加花生)", 70),
    ("香椰冰淇淋(加玉米)", 70),
    ("芒果糯米飯", 80),
    ("香煎香蕉", 60),
    ("香煎香蕉(加冰淇淋)", 70),
]
for name, price in desserts:
    products.append((f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# Write products
for pid, name, price, sort in products:
    ws_products.append([pid, "", name, "", price, False, sort, True, ""])

# Fix category IDs for products
category_map = {cat[0]: cat[1] for cat in categories}
category_ids = {cat[1]: cat[0] for cat in categories}

# Reset and rewrite products with correct category IDs
ws_products.delete_rows(1, ws_products.max_row)
ws_products.append(["productId", "categoryId", "name", "description", "price", "soldOut", "sortOrder", "enabled", "imageUrl"])

# Rebuild products with category IDs
sort_order = 10
products_with_categories = []

# 主食
for dish in main_dishes:
    base_name, base_price, variants = dish
    products_with_categories.append(("cat-main", f"prod-{base_name.lower()[:10].replace(' ', '-')}", base_name, base_price, sort_order))
    sort_order += 10
    for suffix, price_add in variants:
        variant_name = f"{base_name}({suffix})"
        products_with_categories.append(("cat-main", f"prod-{base_name.lower()[:5]}-{suffix[:1]}", variant_name, base_price + price_add, sort_order))
        sort_order += 10

for dish in main_dishes_no_variants:
    name, price = dish
    products_with_categories.append(("cat-main", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 套餐
for name, price in set_meals:
    products_with_categories.append(("cat-set", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 涼拌
for name, price in salads:
    products_with_categories.append(("cat-salad", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 湯品
for name, price in soups:
    products_with_categories.append(("cat-soup", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 茶品
for name, price in teas:
    products_with_categories.append(("cat-tea", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 飲品
for name, price in drinks:
    products_with_categories.append(("cat-drink", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

# 甜點
for name, price in desserts:
    products_with_categories.append(("cat-dessert", f"prod-{name.lower()[:10].replace(' ', '-')}", name, price, sort_order))
    sort_order += 10

for cat_id, pid, name, price, sort in products_with_categories:
    ws_products.append([pid, cat_id, name, "", price, False, sort, True, ""])

# ── ProductOptions (bindings) ──
ws_product_options = wb.create_sheet("ProductOptions")
ws_product_options.append(["productId", "groupId", "sortOrder", "enabled"])

# 主食綁定 加料 (only base dishes, not variants)
base_main_dish_names = [d[0] for d in main_dishes] + [d[0] for d in main_dishes_no_variants]
for dish_name in base_main_dish_names:
    pid = f"prod-{dish_name.lower()[:10].replace(' ', '-')}"
    ws_product_options.append([pid, "opt-protein", 10, True])

# 飲品綁定 冰塊 + 甜度
for drink_name, _ in drinks:
    pid = f"prod-{drink_name.lower()[:10].replace(' ', '-')}"
    ws_product_options.append([pid, "opt-ice", 10, True])
    ws_product_options.append([pid, "opt-sweet", 20, True])

# 茶品綁定 甜度
for tea_name, _ in teas:
    pid = f"prod-{tea_name.lower()[:10].replace(' ', '-')}"
    ws_product_options.append([pid, "opt-sweet", 10, True])

# 甜點綁定 加料 (香椰冰淇淋、香煎香蕉)
for dessert_name in ["香椰冰淇淋", "香煎香蕉"]:
    pid = f"prod-{dessert_name.lower()[:10].replace(' ', '-')}"
    ws_product_options.append([pid, "opt-protein", 10, True])

# ── Save ──
output_path = "outputs/art-village-sheet/藝素村菜單匯入檔.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print(f"  Categories: {len(categories)}")
print(f"  Products: {len(products_with_categories)}")
print(f"  OptionGroups: {len(option_groups)}")
print(f"  OptionItems: {len(option_items)}")
print(f"  ProductOptions bindings: {ws_product_options.max_row - 1}")
